import sys, argparse, re
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from datetime import datetime, date

# =============================================================================
# extract_excel.py — trích xuất file Excel "Resource Allocation" thành SQL import.
#
# ⚙️  BẢO TRÌ: gần như mọi thứ cần chỉnh khi template Excel thay đổi nằm trong
#     khối CONFIG ngay bên dưới (tên sheet, số dòng tiêu đề, map sản phẩm, định
#     dạng ngày, header kỳ vọng). Script đọc theo VỊ TRÍ CỘT nên nếu template
#     chèn/xoá/đổi thứ tự cột thì dữ liệu sẽ vào sai trường — khối EXPECTED_HEADERS
#     sẽ IN CẢNH BÁO để phát hiện sớm (xem phần "WARNINGS" ở cuối output).
#
#     File này là BẢN GỐC DUY NHẤT (được đóng gói vào jar backend và chạy khi
#     POST /api/import/excel). Có thể chạy tay:
#        python extract_excel.py --xlsx <file.xlsx> --out <out.sql>
# =============================================================================

# ─────────────────────────────── CONFIG ─────────────────────────────────────
# Tên sheet chấp nhận (get_sheet khớp được cả prefix số / hoa-thường / khoảng trắng).
SHEETS = {
    'customers':       ['DS Khách hàng'],
    'projects':        ['1.DS Dự án', 'DS Dự án'],
    'employees':       ['2.Tổ chức nhân sự', 'Tổ chức nhân sự'],
    'phases':          ['Config'],
    'departments':     ['3.Tổ chức phòng ban', 'Tổ chức phòng ban'],
    'allocations':     ['4.Phân bổ nguồn lực', 'Phân bổ nguồn lực'],
    'allocations_old': ['Phân bổ nguồn lực_Old'],
    'milestones':      ['5.Deliverable List', 'Deliverable List'],
    'tasks':           ['6.Master Schedule', 'Sheet4'],
}

# Số dòng tiêu đề bỏ qua trước khi tới dữ liệu (mặc định 1).
HEADER_ROWS = {'default': 1, 'milestones': 3, 'tasks': 2}
# Dòng chứa tên cột dùng để kiểm tra header (0-based).
HEADER_CHECK_ROW = {'default': 0, 'milestones': 1, 'tasks': 1}

# Cột phân bổ theo tháng: header dạng chuỗi bắt đầu bằng MONTH_PREFIX và chứa MONTH_CONTAINS (vd 'T1-2026').
MONTH_PREFIX = 'T'
MONTH_CONTAINS = '-20'

# Map tên sản phẩm (sheet Master Schedule) -> mã sản phẩm. CẬP NHẬT khi có sản phẩm mới.
PROD_MAP = {
    'Phần mềm quản lý bệnh viện': 'HIS',
    'Phần mềm quản lý xét nghiệm': 'LIS',
    'Phần mềm quản lý lưu trữ và khai thác bệnh án điện tử': 'EMR',
    'Phần mềm quản lý lưu trữ và khai thác hồ sơ khám sức khỏe CBCS': 'KSK',
    'Phần mềm quản lý trung tâm dữ liệu y tế': 'QLTTDL',
    'Phần mềm quản lý tích hợp đồng bộ dữ liệu': 'THDB',
}
PROD_CODE_SET = {'HIS', 'LIS', 'EMR', 'KSK', 'QLTTDL', 'THDB'}

# Định dạng ngày thử parse (theo thứ tự) và độ dài tối đa để đoán 1 chuỗi là ngày.
DATE_FORMATS = ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y', '%-d/%-m/%Y')
DATE_COERCE_MAXLEN = 12

# Header kỳ vọng (chỉ CẢNH BÁO nếu lệch, không chặn import). Khóa = chỉ số cột, giá trị = từ khóa con.
EXPECTED_HEADERS = {
    'customers':       {1: 'Chủ đầu tư', 2: 'Tên chủ đầu tư', 3: 'Đơn vị thụ hưởng', 5: 'Mã Đề án'},
    'projects':        {0: 'Mã Đề án', 3: 'Mã Sản phẩm', 4: 'Sản phẩm', 5: 'PM', 6: 'Start', 7: 'End'},
    'employees':       {1: 'Nhân sự', 2: 'Account', 3: 'Công ty', 4: 'Trung tâm', 5: 'Role'},
    'phases':          {0: 'Giai đoạn', 1: 'Giai đoạn thành phần'},
    'departments':     {0: 'Trung tâm', 1: 'Bộ phận'},
    'allocations':     {0: 'Dự án', 1: 'Nhân sự', 2: 'Vai trò', 3: 'From', 4: 'To'},
    'allocations_old': {0: 'Dự án', 1: 'Nhân sự', 2: 'Vai trò', 3: 'From', 4: 'To'},
    'milestones':      {1: 'Mã dự án', 3: 'Milestone thành phần', 4: 'Milestone chi tiết', 16: 'Trạng thái'},
    'tasks':           {0: 'Mã dự án', 1: 'Giai đoạn', 2: 'Task', 3: 'Nội dung'},
}
# ──────────────────────────── HẾT CONFIG ────────────────────────────────────

parser = argparse.ArgumentParser(description='Extract Excel to SQL')
parser.add_argument('--xlsx', default=r'C:/Users/Mr.Cua/Downloads/Copy of Resource Allocation_V1.0.xlsx')
parser.add_argument('--out',  default=r'D:/QT_175/database/full_data.sql')
args = parser.parse_args()

XLSX = args.xlsx
OUT  = args.out

WARNINGS = []
def warn(msg):
    WARNINGS.append(msg)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

def parse_date(s):
    """Try to parse date string in multiple formats."""
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(str(s).strip(), fmt).strftime('%Y-%m-%d')
        except: pass
    return None

def esc(v):
    if v is None: return 'NULL'
    if isinstance(v, bool): return '1' if v else '0'
    if isinstance(v, (int, float)): return str(round(v, 6))
    if isinstance(v, (datetime, date)):
        d = v if isinstance(v, date) else v.date()
        return f"'{d.strftime('%Y-%m-%d')}'"
    s = str(v).strip()
    # Try to detect date strings
    if s and ('/' in s or '-' in s) and len(s) <= DATE_COERCE_MAXLEN:
        parsed = parse_date(s)
        if parsed: return f"'{parsed}'"
    s = s.replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"

def _norm_sheet(name):
    """Normalize a sheet name for matching: drop leading 'N.' index, lowercase, collapse spaces."""
    s = re.sub(r'^\s*\d+\s*\.\s*', '', str(name))
    return ' '.join(s.split()).strip().lower()

def get_sheet(wb, *names):
    """Find a sheet by name, tolerant of leading numbered prefixes (e.g. '1.DS Dự án'),
    case and surrounding whitespace. Supports both old (plain) and new (numbered) naming."""
    # 1) exact match
    for n in names:
        if n in wb.sheetnames:
            return wb[n]
    # 2) normalized match (ignore 'N.' prefix / case / extra spaces)
    norm_map = {_norm_sheet(s): s for s in wb.sheetnames}
    for n in names:
        real = norm_map.get(_norm_sheet(n))
        if real is not None:
            return wb[real]
    raise KeyError(f"Không tìm thấy sheet {list(names)}. Các sheet hiện có: {wb.sheetnames}")

def load_rows(key):
    """Lấy các dòng không rỗng của một sheet theo CONFIG. Nếu thiếu sheet -> cảnh báo và trả None
    (bỏ qua nguyên section thay vì làm hỏng cả import)."""
    try:
        ws = get_sheet(wb, *SHEETS[key])
    except KeyError as e:
        warn(f"[{key}] {e} → BỎ QUA section này (dữ liệu bảng liên quan giữ nguyên).")
        return None
    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    check_header(key, rows)
    return rows

def check_header(key, rows):
    """So khớp header thực tế với EXPECTED_HEADERS để phát hiện chèn/đổi thứ tự cột."""
    exp = EXPECTED_HEADERS.get(key)
    if not exp: return
    hr = HEADER_CHECK_ROW.get(key, HEADER_CHECK_ROW['default'])
    if hr >= len(rows):
        warn(f"[{key}] không có dòng tiêu đề (row {hr}).")
        return
    header = rows[hr]
    for col, kw in exp.items():
        actual = header[col] if col < len(header) else None
        if actual is None or kw.lower() not in str(actual).lower():
            warn(f"[{key}] cột {col} kỳ vọng chứa '{kw}' nhưng thấy: {actual!r} "
                 f"→ có thể template đã đổi thứ tự cột, dữ liệu sẽ vào sai trường!")

def data_rows(rows, key):
    """Cắt bỏ các dòng tiêu đề theo HEADER_ROWS."""
    n = HEADER_ROWS.get(key, HEADER_ROWS['default'])
    return rows[n:]

lines = [
    "-- Auto-generated full data from Excel",
    "SET NAMES utf8mb4;",
    "SET FOREIGN_KEY_CHECKS=0;",
    ""
]

# Bộ đếm (khởi tạo trước để summary luôn chạy kể cả khi 1 section bị bỏ qua).
counts = {}
ms_count = 0
task_count = 0

# ── 1. CUSTOMERS ───────────────────────────────────────────────────────────────
rows = load_rows('customers')
if rows is not None:
    lines += ["-- 1. Customers", "DELETE FROM customers;"]
    c = 0
    for r in data_rows(rows, 'customers'):
        inv_code, inv_name, ben_code, ben_name, pg_code = r[1], r[2], r[3], r[4], r[5]
        if not inv_code: continue
        lines.append(
            f"INSERT IGNORE INTO customers (investor_code,investor_name,beneficiary_code,beneficiary_name,project_group_code)"
            f" VALUES ({esc(inv_code)},{esc(inv_name)},{esc(ben_code)},{esc(ben_name)},{esc(pg_code)});"
        )
        c += 1
    counts['customers'] = c
    lines.append("")

# ── 2. PROJECT GROUPS + PRODUCTS ───────────────────────────────────────────────
rows = load_rows('projects')
if rows is not None:
    lines += ["-- 2. Project Groups", "DELETE FROM products;", "DELETE FROM project_groups;"]
    pg_seen = set()
    for r in data_rows(rows, 'projects'):
        pg_code, pg_name, director = r[0], r[1], r[2]
        if not pg_code or pg_code in pg_seen: continue
        pg_seen.add(pg_code)
        lines.append(
            f"INSERT IGNORE INTO project_groups (code,name,director,customer_id)"
            f" VALUES ({esc(pg_code)},{esc(pg_name)},{esc(director)},"
            f"(SELECT id FROM customers WHERE project_group_code={esc(pg_code)} LIMIT 1));"
        )
    lines += ["", "-- Products"]
    c = 0
    for r in data_rows(rows, 'projects'):
        pg_code, prod_code, prod_name = r[0], r[3], r[4]
        pm, start, end_, status, phase = r[5], r[6], r[7], r[8], r[9]
        has_plan = r[10]
        if not prod_code: continue
        hp = '1' if has_plan else '0'
        lines.append(
            f"INSERT IGNORE INTO products (code,name,project_group_id,pm,start_date,end_date,status,current_phase,has_work_plan)"
            f" VALUES ({esc(prod_code)},{esc(prod_name)},"
            f"(SELECT id FROM project_groups WHERE code={esc(pg_code)} LIMIT 1),"
            f"{esc(pm)},{esc(start)},{esc(end_)},{esc(status)},{esc(phase)},{hp});"
        )
        c += 1
    counts['products'] = c
    lines.append("")

# ── 3. EMPLOYEES ───────────────────────────────────────────────────────────────
rows = load_rows('employees')
if rows is not None:
    lines += ["-- 3. Employees", "DELETE FROM employees;"]
    c = 0
    emp_names = {}
    for r in data_rows(rows, 'employees'):
        stt = r[0]
        if not stt or not isinstance(stt, (int, float)): continue
        name, account, company, dept = r[1], r[2], r[3], r[4]
        role, level, contract, work_mode = r[5], r[6], r[7], r[8]
        work_time, payment_type, work_status = r[9], r[10], r[11]
        if not name: continue
        emp_names[name] = emp_names.get(name, 0) + 1
        lines.append(
            f"INSERT IGNORE INTO employees (name,account,company,department,role,level,contract_type,work_mode,work_time,payment_type,work_status)"
            f" VALUES ({esc(name)},{esc(account)},{esc(company)},{esc(dept)},"
            f"{esc(role)},{esc(level)},{esc(contract)},{esc(work_mode)},{esc(work_time)},{esc(payment_type)},{esc(work_status)});"
        )
        c += 1
    counts['employees'] = c
    dups = [n for n, k in emp_names.items() if k > 1]
    if dups:
        warn(f"[employees] TRÙNG TÊN nhân sự: {dups} → phân bổ khớp theo tên (LIMIT 1) có thể gán nhầm người.")
    lines.append("")

# ── 4. PHASES (Config) ─────────────────────────────────────────────────────────
rows = load_rows('phases')
if rows is not None:
    lines += ["-- 4. Phases", "DELETE FROM phases;"]
    sort_order = 0; current_group = None
    for r in data_rows(rows, 'phases'):
        grp, phase_name = r[0], r[1]
        if grp: current_group = grp.strip()
        if not phase_name: continue
        sort_order += 1
        note = r[2] if len(r) > 2 else None
        pn = phase_name.strip().replace('\n', ' ')
        lines.append(
            f"INSERT INTO phases (phase_group,phase_name,note,sort_order)"
            f" VALUES ({esc(current_group)},{esc(pn)},{esc(note)},{sort_order});"
        )
    counts['phases'] = sort_order
    lines.append("")

# ── 5. DEPARTMENTS ─────────────────────────────────────────────────────────────
rows = load_rows('departments')
if rows is not None:
    lines += ["-- 5. Departments", "DELETE FROM departments;"]
    so = 0
    for r in data_rows(rows, 'departments'):
        center, division, manager = r[0], r[1], r[2]
        if not center and not division: continue
        so += 1
        lines.append(
            f"INSERT IGNORE INTO departments (center,division,manager,sort_order)"
            f" VALUES ({esc(center)},{esc(division)},{esc(manager)},{so});"
        )
    counts['departments'] = so
    lines.append("")

# ── 6. RESOURCE ALLOCATIONS + MONTHLY ─────────────────────────────────────────
rows = load_rows('allocations')
if rows is not None:
    header = rows[0]
    month_cols = [(i, h) for i, h in enumerate(header)
                  if h and isinstance(h, str) and h.startswith(MONTH_PREFIX) and MONTH_CONTAINS in h]
    if not month_cols:
        warn("[allocations] không phát hiện cột tháng nào (header dạng 'T?-20??') → monthly_allocations rỗng.")
    lines += ["-- 6. Resource Allocations", "DELETE FROM monthly_allocations;", "DELETE FROM resource_allocations;"]
    c = 0
    for r in data_rows(rows, 'allocations'):
        prod_code, emp_name, role = r[0], r[1], r[2]
        from_date, to_date, pct = r[3], r[4], r[5]
        if not prod_code or not emp_name: continue
        lines.append(
            f"INSERT INTO resource_allocations (product_id,employee_id,role_in_project,from_date,to_date,allocation_percent)"
            f" VALUES ("
            f"(SELECT id FROM products WHERE code={esc(prod_code)} LIMIT 1),"
            f"(SELECT id FROM employees WHERE name={esc(emp_name)} LIMIT 1),"
            f"{esc(role)},{esc(from_date)},{esc(to_date)},{esc(pct)});"
        )
        c += 1
        for col_idx, month_label in month_cols:
            val = r[col_idx] if col_idx < len(r) else None
            if val is not None and isinstance(val, (int, float)):
                lines.append(
                    f"INSERT INTO monthly_allocations (allocation_id,`year_month`,`percent`)"
                    f" VALUES (LAST_INSERT_ID(),{esc(month_label)},{round(float(val),4)});"
                )
    counts['resource_allocations'] = c
    lines.append("")

# ── 7. ALLOCATION HISTORY OLD ──────────────────────────────────────────────────
rows = load_rows('allocations_old')
if rows is not None:
    header = rows[0]
    month_cols_old = [(i, h) for i, h in enumerate(header)
                      if h and isinstance(h, str) and h.startswith(MONTH_PREFIX) and MONTH_CONTAINS in h]
    if not month_cols_old:
        warn("[allocations_old] không phát hiện cột tháng nào → allocation_history_monthly rỗng.")
    lines += ["-- 7. Allocation History (Old)", "DELETE FROM allocation_history_monthly;", "DELETE FROM allocation_history;"]
    c = 0
    for r in data_rows(rows, 'allocations_old'):
        projects, emp_name, role = r[0], r[1], r[2]
        from_date, to_date, pct = r[3], r[4], r[5]
        if not projects or not emp_name: continue
        lines.append(
            f"INSERT INTO allocation_history (projects_text,employee_name,role_in_project,from_date,to_date,allocation_percent)"
            f" VALUES ({esc(projects)},{esc(emp_name)},{esc(role)},{esc(from_date)},{esc(to_date)},{esc(pct)});"
        )
        c += 1
        for col_idx, month_label in month_cols_old:
            val = r[col_idx] if col_idx < len(r) else None
            if val is not None and isinstance(val, (int, float)):
                lines.append(
                    f"INSERT INTO allocation_history_monthly (history_id,`year_month`,`percent`)"
                    f" VALUES (LAST_INSERT_ID(),{esc(month_label)},{round(float(val),4)});"
                )
    counts['allocation_history'] = c
    lines.append("")

# ── 8. MILESTONES (Deliverable List) ──────────────────────────────────────────
# Structure: row0+1+2 = 3 header rows, data from row3
# col[0]=phase, col[1]=prod_code, col[3]=component, col[4]=detail(plan), col[5]=detail(sys)
# col[6]=has_gtel(x), col[7]=has_outsource_gtel(x), col[8]=has_outsource_online(x)
# col[10]=plan_start, col[11]=plan_end, col[12]=remind
# col[13]=actual_start, col[14]=actual_end, col[15]=current_phase, col[16]=status
rows = load_rows('milestones')
if rows is not None:
    lines += ["-- 8. Milestones (Deliverable List)", "DELETE FROM milestones;"]
    cur_phase = None; cur_prod = None
    for r in data_rows(rows, 'milestones'):  # skip 3 header rows
        phase = r[0]; prod_code = r[1]
        component = r[3]; detail = r[4]; detail_sys = r[5]
        has_gtel = 1 if r[6] and str(r[6]).strip().lower() in ('x','1','true') else 0
        has_out_gtel = 1 if r[7] and str(r[7]).strip().lower() in ('x','1','true') else 0
        has_out_online = 1 if r[8] and str(r[8]).strip().lower() in ('x','1','true') else 0
        plan_start = r[10]; plan_end = r[11]; remind = r[12]
        act_start = r[13]; act_end = r[14]; cur_ph = r[15]; status = r[16]

        if phase: cur_phase = phase
        if prod_code: cur_prod = prod_code

        # Only insert rows that have a detail milestone
        if not detail and not detail_sys: continue
        if not cur_prod: continue

        ms_count += 1
        detail_str = detail or detail_sys
        lines.append(
            f"INSERT INTO milestones (product_id,phase,component_milestone,detail_milestone,"
            f"has_gtel,has_outsource_gtel,has_outsource_online,"
            f"plan_start_date,plan_end_date,remind,actual_start_date,actual_end_date,current_phase,status)"
            f" VALUES ("
            f"(SELECT id FROM products WHERE code={esc(cur_prod)} LIMIT 1),"
            f"{esc(cur_phase)},{esc(component)},{esc(detail_str)},"
            f"{has_gtel},{has_out_gtel},{has_out_online},"
            f"{esc(plan_start)},{esc(plan_end)},{esc(remind)},"
            f"{esc(act_start)},{esc(act_end)},{esc(cur_ph)},{esc(status)});"
        )
    counts['milestones'] = ms_count
    if ms_count == 0:
        warn("[milestones] 0 milestone được tạo → kiểm tra sheet 'Deliverable List' (cột chi tiết/mã dự án).")
    lines.append(f"-- {ms_count} milestones inserted")
    lines.append("")

# ── 9. TASKS (Sheet4 - Master Schedule) ───────────────────────────────────────
rows = load_rows('tasks')
if rows is not None:
    lines += ["-- 9. Tasks (Master Schedule - Sheet4)", "DELETE FROM tasks;"]
    # Structure: c0=product_code(only on 1st row) OR phase name, c1=task_no(float), c2=task_name, c3=content, c4=feature_group
    cur_phase = None; cur_prod_code = None
    PROD_CODES = PROD_CODE_SET | set(PROD_MAP.keys())
    unknown_groups = set()
    for r in data_rows(rows, 'tasks'):
        c0 = r[0]; c1 = r[1]; c2 = r[2]; c3 = r[3]
        c4 = r[4] if len(r) > 4 else None

        # Phase group: c0 has text, c1 is None
        if c0 and isinstance(c0, str) and c0.strip() and c1 is None:
            c0s = c0.strip()
            # Check if it's a known product code
            if c0s in PROD_MAP:
                cur_prod_code = PROD_MAP[c0s]
            elif c0s in PROD_CODES:
                cur_prod_code = c0s
            else:
                cur_phase = c0s
            continue

        # Task row: c0=product_code c1=task_no
        if c0 and isinstance(c0, str) and c0.strip() in PROD_CODE_SET:
            cur_prod_code = c0.strip()
            # c1 = task_no, c2 = task_name, c3 = content, c4 = feature_group
            if c1 is not None and isinstance(c1, (int, float)) and c2:
                task_count += 1
                lines.append(
                    f"INSERT INTO tasks (phase_group,product_id,task_no,task_name,feature_group,content)"
                    f" VALUES ({esc(cur_phase)},"
                    f"(SELECT id FROM products WHERE code={esc(cur_prod_code)} LIMIT 1),"
                    f"{int(c1)},{esc(c2)},{esc(c4 or c2)},{esc(c3)});"
                )
            continue

        # Continuation row: c0=None, c1=task_no
        if c0 is None and c1 is not None and isinstance(c1, (int, float)) and c2 and cur_prod_code:
            task_count += 1
            lines.append(
                f"INSERT INTO tasks (phase_group,product_id,task_no,task_name,feature_group,content)"
                f" VALUES ({esc(cur_phase)},"
                f"(SELECT id FROM products WHERE code={esc(cur_prod_code)} LIMIT 1),"
                f"{int(c1)},{esc(c2)},{esc(c4 or c2)},{esc(c3)});"
            )
    counts['tasks'] = task_count
    if task_count == 0:
        warn("[tasks] 0 task được tạo → kiểm tra sheet Master Schedule và PROD_MAP (tên sản phẩm mới?).")
    lines += [f"-- {task_count} tasks inserted", ""]

lines += ["SET FOREIGN_KEY_CHECKS=1;", ""]

with open(OUT, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f"Generated: {OUT}")
print(f"Total SQL lines: {len(lines)}")
print(f"Milestones: {ms_count} | Tasks: {task_count}")
print("Row counts: " + ", ".join(f"{k}={v}" for k, v in counts.items()))
if WARNINGS:
    print(f"WARNINGS ({len(WARNINGS)}):")
    for w in WARNINGS:
        print("  ⚠ " + w)
else:
    print("WARNINGS: none")
